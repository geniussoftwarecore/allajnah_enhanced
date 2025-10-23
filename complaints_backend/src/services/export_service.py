# type: ignore
import os
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from src.models.complaint import Complaint, Payment, User, Subscription, Role
from src.database.db import db

class ExportService:
    """Service for exporting data to Excel with Arabic RTL support"""
    
    @staticmethod
    def _apply_rtl_formatting(file_path):
        """Apply RTL formatting and styling to Excel file"""
        workbook = load_workbook(file_path)
        worksheet = workbook.active
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
        
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
        
        worksheet.sheet_view.rightToLeft = True
        
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        workbook.save(file_path)
        return file_path
    
    @staticmethod
    def export_complaints_to_excel(filters=None):
        """
        Export complaints to Excel with filters
        
        Args:
            filters: dict with optional keys:
                - start_date: datetime
                - end_date: datetime
                - status_id: int
                - category_id: int
                - trader_id: str
        
        Returns:
            str: File path to generated Excel file
        """
        query = Complaint.query
        
        if filters:
            if filters.get('start_date'):
                query = query.filter(Complaint.submitted_at >= filters['start_date'])
            if filters.get('end_date'):
                query = query.filter(Complaint.submitted_at <= filters['end_date'])
            if filters.get('status_id'):
                query = query.filter(Complaint.status_id == filters['status_id'])
            if filters.get('category_id'):
                query = query.filter(Complaint.category_id == filters['category_id'])
            if filters.get('trader_id'):
                query = query.filter(Complaint.trader_id == filters['trader_id'])
        
        complaints = query.all()
        
        data = []
        for complaint in complaints:
            data.append({
                'رقم الشكوى': complaint.complaint_id,
                'العنوان': complaint.title,
                'الوصف': complaint.description,
                'اسم التاجر': complaint.trader.full_name if complaint.trader else '',
                'الفئة': complaint.category.category_name if complaint.category else '',
                'الحالة': complaint.status.status_name if complaint.status else '',
                'الأولوية': complaint.priority,
                'تاريخ التقديم': complaint.submitted_at.strftime('%Y-%m-%d %H:%M') if complaint.submitted_at else '',
                'آخر تحديث': complaint.last_updated_at.strftime('%Y-%m-%d %H:%M') if complaint.last_updated_at else '',
                'المسند إليه': complaint.assigned_committee_member.full_name if complaint.assigned_committee_member else '',
                'تفاصيل الحل': complaint.resolution_details or '',
                'تاريخ الإغلاق': complaint.closed_at.strftime('%Y-%m-%d %H:%M') if complaint.closed_at else '',
                'عدد المرفقات': len(complaint.attachments),
                'عدد التعليقات': len(complaint.comments)
            })
        
        df = pd.DataFrame(data)
        
        exports_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'exports')
        os.makedirs(exports_dir, exist_ok=True)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'complaints_export_{timestamp}.xlsx'
        file_path = os.path.join(exports_dir, filename)
        
        df.to_excel(file_path, index=False, engine='openpyxl')
        
        return ExportService._apply_rtl_formatting(file_path)
    
    @staticmethod
    def export_payments_to_excel(filters=None):
        """
        Export payments to Excel with filters
        
        Args:
            filters: dict with optional keys:
                - start_date: datetime
                - end_date: datetime
                - status: str
                - user_id: str
        
        Returns:
            str: File path to generated Excel file
        """
        query = Payment.query
        
        if filters:
            if filters.get('start_date'):
                query = query.filter(Payment.payment_date >= filters['start_date'])
            if filters.get('end_date'):
                query = query.filter(Payment.payment_date <= filters['end_date'])
            if filters.get('status'):
                query = query.filter(Payment.status == filters['status'])
            if filters.get('user_id'):
                query = query.filter(Payment.user_id == filters['user_id'])
        
        payments = query.all()
        
        data = []
        for payment in payments:
            data.append({
                'رقم الدفع': payment.payment_id,
                'اسم المستخدم': payment.user.full_name if payment.user else '',
                'اسم المرسل': payment.sender_name,
                'هاتف المرسل': payment.sender_phone,
                'طريقة الدفع': payment.payment_method.name if payment.payment_method else '',
                'المبلغ': payment.amount,
                'العملة': payment.currency,
                'رقم المعاملة': payment.transaction_reference or '',
                'تاريخ الدفع': payment.payment_date.strftime('%Y-%m-%d %H:%M') if payment.payment_date else '',
                'الحالة': payment.status,
                'المراجع': payment.reviewed_by.full_name if payment.reviewed_by else '',
                'ملاحظات المراجعة': payment.review_notes or '',
                'تاريخ المراجعة': payment.reviewed_at.strftime('%Y-%m-%d %H:%M') if payment.reviewed_at else '',
                'تاريخ الإنشاء': payment.created_at.strftime('%Y-%m-%d %H:%M') if payment.created_at else ''
            })
        
        df = pd.DataFrame(data)
        
        exports_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'exports')
        os.makedirs(exports_dir, exist_ok=True)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'payments_export_{timestamp}.xlsx'
        file_path = os.path.join(exports_dir, filename)
        
        df.to_excel(file_path, index=False, engine='openpyxl')
        
        return ExportService._apply_rtl_formatting(file_path)
    
    @staticmethod
    def export_users_to_excel(role_filter=None):
        """
        Export users to Excel with optional role filter
        
        Args:
            role_filter: str or list of role names to filter by
        
        Returns:
            str: File path to generated Excel file
        """
        query = User.query
        
        if role_filter:
            if isinstance(role_filter, str):
                role_filter = [role_filter]
            query = query.join(Role).filter(Role.role_name.in_(role_filter))
        
        users = query.all()
        
        data = []
        for user in users:
            data.append({
                'رقم المستخدم': user.user_id,
                'اسم المستخدم': user.username,
                'البريد الإلكتروني': user.email,
                'الاسم الكامل': user.full_name,
                'رقم الهاتف': user.phone_number or '',
                'العنوان': user.address or '',
                'الدور': user.role.role_name if user.role else '',
                'نشط': 'نعم' if user.is_active else 'لا',
                'المصادقة الثنائية': 'مفعلة' if user.two_factor_enabled else 'غير مفعلة',
                'تاريخ الإنشاء': user.created_at.strftime('%Y-%m-%d %H:%M') if user.created_at else '',
                'آخر تحديث': user.updated_at.strftime('%Y-%m-%d %H:%M') if user.updated_at else ''
            })
        
        df = pd.DataFrame(data)
        
        exports_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'exports')
        os.makedirs(exports_dir, exist_ok=True)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'users_export_{timestamp}.xlsx'
        file_path = os.path.join(exports_dir, filename)
        
        df.to_excel(file_path, index=False, engine='openpyxl')
        
        return ExportService._apply_rtl_formatting(file_path)
    
    @staticmethod
    def export_subscriptions_to_excel():
        """
        Export subscriptions to Excel
        
        Returns:
            str: File path to generated Excel file
        """
        subscriptions = Subscription.query.all()
        
        data = []
        for subscription in subscriptions:
            data.append({
                'رقم الاشتراك': subscription.subscription_id,
                'اسم المستخدم': subscription.user.full_name if subscription.user else '',
                'البريد الإلكتروني': subscription.user.email if subscription.user else '',
                'الخطة': subscription.plan,
                'تاريخ البدء': subscription.start_date.strftime('%Y-%m-%d') if subscription.start_date else '',
                'تاريخ الانتهاء': subscription.end_date.strftime('%Y-%m-%d') if subscription.end_date else '',
                'الحالة': subscription.status,
                'تجديد': 'نعم' if subscription.is_renewal else 'لا',
                'فترة سماح': 'مفعلة' if subscription.grace_period_enabled else 'غير مفعلة',
                'إشعار 14 يوم': 'تم' if subscription.notified_14d else 'لم يتم',
                'إشعار 7 أيام': 'تم' if subscription.notified_7d else 'لم يتم',
                'إشعار 3 أيام': 'تم' if subscription.notified_3d else 'لم يتم',
                'تاريخ الإنشاء': subscription.created_at.strftime('%Y-%m-%d %H:%M') if subscription.created_at else ''
            })
        
        df = pd.DataFrame(data)
        
        exports_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'exports')
        os.makedirs(exports_dir, exist_ok=True)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'subscriptions_export_{timestamp}.xlsx'
        file_path = os.path.join(exports_dir, filename)
        
        df.to_excel(file_path, index=False, engine='openpyxl')
        
        return ExportService._apply_rtl_formatting(file_path)
